import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

# Set up WebDriver and Excel file
driver = webdriver.Chrome(options=webdriver.ChromeOptions().add_argument("--start-maximized"))
excel_file = r"C:\Users\parve\OneDrive\Desktop\4Beats-Task1\output.xlsx"
sheet_name = datetime.datetime.now().strftime("%A")
wb = load_workbook(excel_file)
sheet = wb[sheet_name]

# Process each keyword
for row_num, (keyword,) in enumerate(sheet.iter_rows(min_row=3, values_only=True, min_col=3, max_col=3), start=3):
    if keyword:
        keyword = keyword.strip()
        driver.get("https://www.google.com")
        driver.find_element("name", "q").send_keys(keyword)
        time.sleep(2)

        # Get suggestions
        suggestions = [s.find_element(By.XPATH, ".//div").text for s in driver.find_elements(By.CSS_SELECTOR, ".erkvQe li")]
        
        if suggestions:
            longest = max(suggestions, key=len)
            shortest = min(suggestions, key=len)
            sheet.cell(row=row_num, column=4, value=longest)
            sheet.cell(row=row_num, column=5, value=shortest)

# Save changes and print status
wb.save(excel_file)
print("Longest and shortest suggestions added to the columns.")

# Close the browser
driver.quit()
